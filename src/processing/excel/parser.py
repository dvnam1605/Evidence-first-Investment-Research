"""Synchronous openpyxl XLSX extractor."""

from __future__ import annotations

import hashlib
import io
from datetime import date, datetime, time
from decimal import Decimal
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.processing.errors import ExcelParseError
from src.processing.excel.models import (
    ExcelCell,
    ExcelCellValueKind,
    ExcelSheet,
    ExcelSheetVisibility,
    ParsedWorkbook,
)

PARSER_NAME = "openpyxl"


def _openpyxl_version() -> str:
    try:
        return version("openpyxl")
    except PackageNotFoundError:
        return "unknown"


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _value_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _merge_map(sheet: Worksheet) -> dict[tuple[int, int], str]:
    """Map every cell coordinate inside a merge to its range string (e.g. A1:B2)."""
    mapping: dict[tuple[int, int], str] = {}
    for merged in sheet.merged_cells.ranges:
        range_str = str(merged)
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                mapping[(row, col)] = range_str
    return mapping


def _sheet_visibility(sheet: Worksheet) -> ExcelSheetVisibility:
    state = str(getattr(sheet, "sheet_state", ExcelSheetVisibility.VISIBLE.value))
    try:
        return ExcelSheetVisibility(state)
    except ValueError:
        return ExcelSheetVisibility.VISIBLE


def _cell_from_openpyxl(
    *,
    sheet_name: str,
    row_idx: int,
    col_idx: int,
    address: str,
    value: Any,
    number_format: str,
    data_type: str | None,
    merge_range: str | None,
) -> ExcelCell:
    is_merged = merge_range is not None
    if data_type == "f" or (isinstance(value, str) and value.startswith("=")):
        formula = str(value)
        return ExcelCell(
            sheet_name=sheet_name,
            row=row_idx,
            column=col_idx,
            address=address,
            value_kind=ExcelCellValueKind.FORMULA,
            value=formula,
            value_text=formula,
            number_format=number_format,
            formula_text=formula,
            # data_only=False: we deliberately do not evaluate formulas.
            cached_value=None,
            is_merged=is_merged,
            merge_range=merge_range,
        )
    if value is None:
        return ExcelCell(
            sheet_name=sheet_name,
            row=row_idx,
            column=col_idx,
            address=address,
            value_kind=ExcelCellValueKind.EMPTY,
            value=None,
            value_text=None,
            number_format=number_format,
            formula_text=None,
            cached_value=None,
            is_merged=is_merged,
            merge_range=merge_range,
        )
    return ExcelCell(
        sheet_name=sheet_name,
        row=row_idx,
        column=col_idx,
        address=address,
        value_kind=ExcelCellValueKind.SCALAR,
        value=value,
        value_text=_value_text(value),
        number_format=number_format,
        formula_text=None,
        cached_value=None,
        is_merged=is_merged,
        merge_range=merge_range,
    )


def _iter_sheet_cells(sheet: Worksheet) -> list[ExcelCell]:
    merges = _merge_map(sheet)
    cells: list[ExcelCell] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                # Non-anchor cells in a merge: still record coordinates for provenance.
                if cell.row is None or cell.column is None:
                    continue
                row_idx = int(cell.row)
                col_idx = int(cell.column)
                merge_range = merges.get((row_idx, col_idx))
                cells.append(
                    _cell_from_openpyxl(
                        sheet_name=sheet.title,
                        row_idx=row_idx,
                        col_idx=col_idx,
                        address=f"{get_column_letter(col_idx)}{row_idx}",
                        value=None,
                        number_format="General",
                        data_type=None,
                        merge_range=merge_range,
                    )
                )
                continue

            if not isinstance(cell, Cell):
                continue

            row_idx = int(cell.row)
            col_idx = int(cell.column)
            merge_range = merges.get((row_idx, col_idx))
            value = cell.value
            # Skip completely empty non-merged cells to keep payloads focused.
            if value is None and merge_range is None:
                continue

            cells.append(
                _cell_from_openpyxl(
                    sheet_name=sheet.title,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    address=str(cell.coordinate),
                    value=value,
                    number_format=str(cell.number_format or "General"),
                    data_type=str(cell.data_type) if cell.data_type is not None else None,
                    merge_range=merge_range,
                )
            )
    return cells


def _workbook_metadata(workbook: Workbook) -> dict[str, Any]:
    props = workbook.properties
    meta: dict[str, Any] = {}
    if props is None:
        return meta
    for key in ("title", "creator", "subject", "description", "category"):
        value = getattr(props, key, None)
        if value not in (None, ""):
            meta[key] = value
    return meta


class OpenpyxlExcelParser:
    """Synchronous XLSX extractor used by the async ExcelParser service."""

    def parse_path(
        self,
        path: Path,
        *,
        artifact_id: UUID | None = None,
    ) -> ParsedWorkbook:
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")
        return self.parse_bytes(
            path.read_bytes(),
            artifact_id=artifact_id,
            source_label=str(path),
        )

    def parse_bytes(
        self,
        data: bytes,
        *,
        artifact_id: UUID | None = None,
        source_label: str | None = "bytes",
    ) -> ParsedWorkbook:
        if not data:
            raise ExcelParseError("Excel bytes must not be empty")

        source_sha256 = _sha256(data)
        parser_version = _openpyxl_version()
        try:
            workbook = load_workbook(
                filename=io.BytesIO(data),
                data_only=False,
                read_only=False,
            )
        except ExcelParseError:
            raise
        except Exception as exc:  # noqa: BLE001 - openpyxl raises varied types
            raise ExcelParseError(f"Failed to open XLSX bytes: {exc}") from exc

        try:
            if not workbook.sheetnames:
                raise ExcelParseError("Workbook has zero worksheets")

            sheets: list[ExcelSheet] = []
            for index, name in enumerate(workbook.sheetnames):
                worksheet = workbook[name]
                if not isinstance(worksheet, Worksheet):
                    # Charts/macros sheets are skipped for MVP cell extraction.
                    continue
                merged_ranges = tuple(str(r) for r in worksheet.merged_cells.ranges)
                sheets.append(
                    ExcelSheet(
                        name=name,
                        index=index,
                        cells=tuple(_iter_sheet_cells(worksheet)),
                        visibility=_sheet_visibility(worksheet),
                        merged_ranges=merged_ranges,
                    )
                )

            if not sheets:
                raise ExcelParseError("Workbook has no extractable worksheets")

            return ParsedWorkbook(
                sheets=tuple(sheets),
                source_sha256=source_sha256,
                parser_name=PARSER_NAME,
                parser_version=parser_version,
                source_label=source_label,
                artifact_id=artifact_id,
                metadata=_workbook_metadata(workbook),
            )
        finally:
            workbook.close()
