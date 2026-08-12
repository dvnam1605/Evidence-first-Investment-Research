"""Unit tests for DOC-11 table extraction abstraction."""

from __future__ import annotations

import io
import json
from pathlib import Path
from unittest.mock import MagicMock, patch
from uuid import uuid4

import fitz
import pytest
from openpyxl import Workbook
from src.domain.document_block import BoundingBox
from src.processing.errors import PDFParseError
from src.processing.excel.parser import OpenpyxlExcelParser
from src.processing.tables import (
    ExcelGridTableExtractor,
    ExtractedCell,
    ExtractedRow,
    ExtractedTable,
    PassthroughTableReconstructor,
    PyMuPDFTableExtractor,
    TableExtractionContext,
    TableExtractionService,
    TableExtractionStatus,
)
from src.processing.tables.detector import region_from_bbox

TABLE_FIXTURES = Path(__file__).resolve().parents[2] / "fixtures" / "processing" / "tables"
FPT_REAL = TABLE_FIXTURES / "fpt_real"
BBOX_TOL = 0.75


def _load_fixture(name: str) -> tuple[bytes, dict]:
    pdf_path = TABLE_FIXTURES / f"{name}.pdf"
    exp_path = TABLE_FIXTURES / f"{name}.expected.json"
    assert pdf_path.is_file(), f"missing fixture PDF: {pdf_path}"
    assert exp_path.is_file(), f"missing expected JSON: {exp_path}"
    expected = json.loads(exp_path.read_text(encoding="utf-8"))
    return pdf_path.read_bytes(), expected


def _matrix(table: ExtractedTable) -> list[list[str]]:
    return [[cell.raw_text for cell in row.cells] for row in table.rows]


def _assert_bbox_close(
    actual: BoundingBox | None,
    expected: list[float] | None,
    *,
    label: str,
) -> None:
    if expected is None:
        assert actual is None, label
        return
    assert actual is not None, label
    for got, want, axis in (
        (actual.x0, expected[0], "x0"),
        (actual.y0, expected[1], "y0"),
        (actual.x1, expected[2], "x1"),
        (actual.y1, expected[3], "y1"),
    ):
        assert abs(got - want) <= BBOX_TOL, f"{label}.{axis}: {got} vs {want}"


def _assert_table_matches_expected(table: ExtractedTable, expected: dict) -> None:
    assert table.page == expected["page"]
    assert table.table_index == expected["table_index"]
    want_cells: list[list[str]] = expected["cells"]
    got_cells = _matrix(table)
    assert len(got_cells) == len(want_cells), "row count mismatch"
    for r, (got_row, want_row) in enumerate(zip(got_cells, want_cells, strict=True)):
        assert len(got_row) == len(want_row), f"col count mismatch at row {r}"
        for c, (got, want) in enumerate(zip(got_row, want_row, strict=True)):
            assert got == want, f"cell[{r}][{c}]: {got!r} != {want!r}"
            cell = table.rows[r].cells[c]
            assert cell.row == r
            assert cell.column == c
            assert cell.page == table.page
            assert cell.table_id == table.table_id

    _assert_bbox_close(table.bbox, expected.get("bbox"), label="table.bbox")
    cell_bboxes = expected.get("cell_bboxes")
    if cell_bboxes is not None:
        for r, row in enumerate(table.rows):
            for c, cell in enumerate(row.cells):
                _assert_bbox_close(
                    cell.bbox,
                    cell_bboxes[r][c],
                    label=f"cell_bbox[{r}][{c}]",
                )


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "IS"
    ws["A1"] = "Chỉ tiêu"
    ws["B1"] = "Kỳ này"
    ws["A2"] = "Doanh thu thuần"
    ws["B2"] = 1000
    ws["A3"] = "Giá vốn"
    ws["B3"] = 400
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_with_merges_and_formats() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BS"
    ws.merge_cells("A1:B1")
    ws["A1"] = "Bảng cân đối kế toán"
    ws["A2"] = "Tiền và tương đương tiền"
    ws["B2"] = "(1,234.50)"
    ws["A3"] = "Lỗ trong kỳ"
    ws["B3"] = "-50"
    ws["A4"] = "Số dư EU"
    ws["B4"] = "1.000,25"
    ws["A5"] = ""
    ws["B5"] = None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extracted_cell_retains_coordinates_and_table_id() -> None:
    cell = ExtractedCell(
        raw_text="1000",
        row=1,
        column=2,
        page=3,
        table_id="doc:p3:t0",
        bbox=None,
    )
    assert cell.raw_text == "1000"
    assert cell.row == 1
    assert cell.column == 2
    assert cell.page == 3
    assert cell.table_id == "doc:p3:t0"


def test_excel_grid_extractor_preserves_raw_text_and_coords() -> None:
    workbook = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes())
    result = ExcelGridTableExtractor().extract(workbook)
    assert len(result.tables) == 1
    table = result.tables[0]
    assert table.source_label == "IS"
    assert table.page == 1
    assert _matrix(table) == [
        ["Chỉ tiêu", "Kỳ này"],
        ["Doanh thu thuần", "1000"],
        ["Giá vốn", "400"],
    ]
    assert table.rows[1].cells[1].row == 1
    assert table.rows[1].cells[1].column == 1
    assert table.rows[1].cells[1].page == 1
    assert table.rows[1].cells[1].table_id == table.table_id
    assert table.status is TableExtractionStatus.OK
    assert table.confidence is not None
    assert table.confidence < 1.0


def test_excel_merged_continuation_and_locale_raw_values() -> None:
    workbook = OpenpyxlExcelParser().parse_bytes(_xlsx_with_merges_and_formats())
    result = ExcelGridTableExtractor().extract(workbook)
    table = result.tables[0]
    assert _matrix(table) == [
        ["Bảng cân đối kế toán", ""],
        ["Tiền và tương đương tiền", "(1,234.50)"],
        ["Lỗ trong kỳ", "-50"],
        ["Số dư EU", "1.000,25"],
    ]
    assert table.rows[0].cells[1].is_merged_continuation is True
    assert table.rows[0].cells[0].is_merged_continuation is False


def test_passthrough_reconstructor_never_discards_raw() -> None:
    table = ExtractedTable(
        page=1,
        table_id="t1",
        rows=(
            ExtractedRow(
                row_index=0,
                cells=(
                    ExtractedCell(
                        raw_text="a", row=0, column=0, page=1, table_id="t1"
                    ),
                ),
            ),
        ),
        extractor_name="test",
    )
    snap = PassthroughTableReconstructor().reconstruct(table)
    assert snap.raw is table
    assert snap.raw.rows[0].cells[0].raw_text == "a"


def test_region_from_bbox_defaults_uncertain() -> None:
    region = region_from_bbox(page=2, x0=1, y0=2, x1=10, y1=20)
    assert region.page == 2
    assert region.bbox == BoundingBox(1, 2, 10, 20)
    assert region.confidence is None
    assert region.status is TableExtractionStatus.NEEDS_REVIEW


def test_pdf_extractor_no_table_page_returns_empty_result() -> None:
    doc = fitz.open()
    try:
        page = doc.new_page()
        page.insert_text((72, 72), "Not a real table")
        pdf_bytes = doc.tobytes()
    finally:
        doc.close()

    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert result.tables == ()
    assert result.page_issues == ()


def test_pdf_rejects_html_and_non_pdf() -> None:
    extractor = PyMuPDFTableExtractor()
    with pytest.raises(PDFParseError, match="HTML|not a PDF|magic"):
        extractor.extract(b"<html>not a PDF</html>")
    with pytest.raises(PDFParseError, match="empty"):
        extractor.extract(b"")
    with pytest.raises(PDFParseError, match="magic"):
        extractor.extract(b"PK\x03\x04not-a-pdf")


def test_pdf_exact_vi_grid_matrix_and_provenance() -> None:
    pdf_bytes, expected = _load_fixture("exact_vi_grid")
    doc_id = uuid4()
    artifact_id = uuid4()
    result = PyMuPDFTableExtractor().extract(
        pdf_bytes,
        context=TableExtractionContext(document_id=doc_id, artifact_id=artifact_id),
    )
    assert len(result.tables) == 1
    table = result.tables[0]
    _assert_table_matches_expected(table, expected["tables"][0])
    assert table.document_id == doc_id
    assert table.artifact_id == artifact_id
    assert table.table_id.startswith(str(doc_id))
    assert table.confidence is not None
    assert table.confidence < 1.0
    # Full Vietnamese labels and risk-value placements are exact, not any(...).
    assert table.rows[0].cells[0].raw_text == "Chỉ tiêu"
    assert table.rows[1].cells[0].raw_text == "Doanh thu thuần"
    assert table.rows[1].cells[1].raw_text == "(1,234.50)"
    assert table.rows[1].cells[2].raw_text == "1.000,25"
    assert table.rows[2].cells[0].raw_text == "Giá vốn hàng bán"
    assert table.rows[3].cells[0].raw_text == "Lợi nhuận gộp"
    assert table.rows[3].cells[1].raw_text == ""
    assert table.rows[3].cells[2].raw_text == ""


def test_pdf_multiline_vietnamese_label_exact_matrix() -> None:
    pdf_bytes, expected = _load_fixture("multiline_vi_label")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert len(result.tables) == 1
    _assert_table_matches_expected(result.tables[0], expected["tables"][0])
    assert result.tables[0].rows[1].cells[0].raw_text == (
        "Doanh thu thuần\ntừ hợp đồng\nvới khách hàng"
    )


def test_pdf_narrative_adjacent_keeps_single_exact_table() -> None:
    pdf_bytes, expected = _load_fixture("narrative_adjacent")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert len(result.tables) == expected["expected_table_count"] == 1
    _assert_table_matches_expected(result.tables[0], expected["tables"][0])
    # Narrative must not invent extra tables or scramble the grid.
    assert "Ban lãnh đạo" not in "".join(
        cell.raw_text for row in result.tables[0].rows for cell in row.cells
    )


def test_pdf_large_grid_exact_row_column_relationships() -> None:
    pdf_bytes, expected = _load_fixture("large_8x4")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert len(result.tables) == 1
    table = result.tables[0]
    _assert_table_matches_expected(table, expected["tables"][0])
    assert table.row_count == 8
    assert table.column_count == 4
    assert table.rows[7].cells[3].raw_text == "R7C3"
    assert table.rows[3].cells[1].raw_text == "R3C1"


def test_pdf_page_spanning_tables_are_page_local_exact() -> None:
    pdf_bytes, expected = _load_fixture("page_spanning")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert len(result.tables) == expected["expected_table_count"] == 2
    for table, want in zip(result.tables, expected["tables"], strict=True):
        _assert_table_matches_expected(table, want)
    assert result.tables[0].page == 1
    assert result.tables[1].page == 2
    assert result.tables[0].rows[1].cells[0].raw_text == "Doanh thu thuần"
    assert result.tables[1].rows[1].cells[0].raw_text == "Lợi nhuận gộp"
    assert result.tables[1].rows[2].cells[0].raw_text == "Lợi nhuận sau thuế"


def test_pdf_false_positive_layout_yields_no_tables() -> None:
    pdf_bytes, expected = _load_fixture("false_positive_layout")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert result.tables == ()
    assert expected["expected_table_count"] == 0
    assert expected["tables"] == []


def test_pdf_two_tables_exact_matrices_and_order() -> None:
    pdf_bytes, expected = _load_fixture("two_tables_page")
    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert len(result.tables) == expected["expected_table_count"] == 2
    for table, want in zip(result.tables, expected["tables"], strict=True):
        _assert_table_matches_expected(table, want)
    assert result.tables[0].bbox is not None
    assert result.tables[1].bbox is not None
    assert result.tables[0].bbox.y0 <= result.tables[1].bbox.y0
    assert result.tables[0].table_id != result.tables[1].table_id


def test_find_tables_failure_emits_page_issue() -> None:
    doc = fitz.open()
    try:
        doc.new_page()
        pdf_bytes = doc.tobytes()
    finally:
        doc.close()

    def boom(self: object) -> object:  # noqa: ARG001
        raise RuntimeError("simulated detector failure")

    with patch.object(fitz.Page, "find_tables", boom):
        result = PyMuPDFTableExtractor().extract(pdf_bytes)

    assert result.tables == ()
    assert len(result.page_issues) == 1
    issue = result.page_issues[0]
    assert issue.page == 1
    assert issue.reason == "find_tables_failed"
    assert issue.status is TableExtractionStatus.NEEDS_REVIEW
    assert result.needs_review is True


def test_empty_grid_table_needs_review() -> None:
    extractor = PyMuPDFTableExtractor()
    fake_table = MagicMock()
    fake_table.bbox = (10, 10, 100, 100)
    fake_table.extract.return_value = []
    fake_table.rows = []
    ctx = TableExtractionContext(source_label="unit")
    table = extractor._convert_pymupdf_table(
        fake_table,
        page_number=1,
        table_index=0,
        extractor_version="test",
        context=ctx,
    )
    assert table.status is TableExtractionStatus.NEEDS_REVIEW
    assert "empty_grid" in table.warnings
    assert table.confidence is not None
    assert table.confidence < 0.5


def test_merged_slot_none_cell_flagged() -> None:
    extractor = PyMuPDFTableExtractor()
    fake_row = MagicMock()
    fake_row.cells = [(0, 0, 50, 20), None]
    fake_table = MagicMock()
    fake_table.bbox = (0, 0, 100, 40)
    fake_table.extract.return_value = [["Header", ""], ["Value", "10"]]
    fake_table.rows = [fake_row, MagicMock(cells=[(0, 20, 50, 40), (50, 20, 100, 40)])]
    table = extractor._convert_pymupdf_table(
        fake_table,
        page_number=2,
        table_index=0,
        extractor_version="test",
        context=TableExtractionContext(source_label="merge"),
    )
    assert table.rows[0].cells[1].is_merged_continuation is True
    assert table.rows[0].cells[1].raw_text == ""
    assert table.rows[0].cells[0].bbox is not None
    assert table.rows[0].cells[0].bbox.x1 == 50


def test_table_extraction_service_excel() -> None:
    workbook = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes())
    result = TableExtractionService().extract_workbook(workbook)
    assert result.tables[0].extractor_name == "excel_grid"
    assert result.tables[0].rows[2].cells[0].raw_text == "Giá vốn"


def test_provenance_differs_across_documents() -> None:
    pdf_bytes, _ = _load_fixture("exact_vi_grid")
    a = uuid4()
    b = uuid4()
    ta = PyMuPDFTableExtractor().extract(
        pdf_bytes, context=TableExtractionContext(document_id=a)
    ).tables[0]
    tb = PyMuPDFTableExtractor().extract(
        pdf_bytes, context=TableExtractionContext(document_id=b)
    ).tables[0]
    assert _matrix(ta) == _matrix(tb)
    assert ta.page == tb.page
    assert ta.table_index == tb.table_index
    assert ta.table_id != tb.table_id
    assert ta.document_id != tb.document_id


def _image_only_pdf_bytes() -> bytes:
    doc = fitz.open()
    try:
        page = doc.new_page()
        pix = fitz.Pixmap(fitz.csRGB, fitz.IRect(0, 0, 240, 120), 0)
        page.insert_image(fitz.Rect(72, 72, 312, 192), pixmap=pix)
        return doc.tobytes()
    finally:
        doc.close()


def test_image_page_without_detected_tables_needs_review() -> None:
    result = PyMuPDFTableExtractor().extract(_image_only_pdf_bytes())
    assert result.tables == ()
    assert result.needs_review is True
    assert result.page_issues[0].reason == "no_tables_detected"
    assert result.page_issues[0].status is TableExtractionStatus.NEEDS_REVIEW


def test_raster_fpt_financial_pdf_zero_tables_is_reviewable() -> None:
    pdf_path = next(
        FPT_REAL.glob("*Consolidated_Financial_Statements_for_Q22026*.pdf"),
        None,
    )
    if pdf_path is None or not pdf_path.is_file():
        pytest.skip("FPT Q2 2026 consolidated PDF not exported under fpt_real/")
    src = fitz.open(pdf_path)
    try:
        sliced = fitz.open()
        try:
            sliced.insert_pdf(src, from_page=0, to_page=min(2, src.page_count - 1))
            pdf_bytes = sliced.tobytes()
        finally:
            sliced.close()
    finally:
        src.close()

    result = PyMuPDFTableExtractor().extract(pdf_bytes)
    assert result.tables == ()
    assert result.needs_review is True
    assert any(issue.reason == "no_tables_detected" for issue in result.page_issues)


def test_fpt_real_fixtures_exact_grids() -> None:
    """Assert exact grids only when a PDF has a paired expected.json."""
    pairs = [
        pdf_path
        for pdf_path in sorted(FPT_REAL.glob("*.pdf"))
        if pdf_path.with_suffix(".expected.json").is_file()
    ]
    if not pairs:
        pytest.skip(
            "Raster FPT PDFs have no native find_tables grids. "
            "Use tests/fixtures/processing/tables/fpt_raw/*.json "
            "(see test_fpt_raw_reconstruction.py)."
        )
    for pdf_path in pairs:
        expected = json.loads(
            pdf_path.with_suffix(".expected.json").read_text(encoding="utf-8")
        )
        result = PyMuPDFTableExtractor().extract(pdf_path.read_bytes())
        assert len(result.tables) == len(expected["tables"])
        for table, want in zip(result.tables, expected["tables"], strict=True):
            _assert_table_matches_expected(table, want)
