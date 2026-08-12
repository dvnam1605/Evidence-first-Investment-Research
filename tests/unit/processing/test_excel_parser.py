"""Unit tests for DOC-08 Excel / XLSX parser."""

from __future__ import annotations

import io
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from src.processing.errors import ExcelParseError
from src.processing.excel.models import ExcelCellValueKind, ExcelSheetVisibility
from src.processing.excel.parser import OpenpyxlExcelParser
from src.processing.excel.service import ExcelParser


def _xlsx_bytes(build: Callable[[Workbook], None]) -> bytes:
    workbook = Workbook()
    build(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_sheet_names_and_cell_coordinates() -> None:
    def build(wb: Workbook) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = "BS"
        ws["A1"] = "Doanh thu thuần"
        ws["B2"] = 1_234_567
        ws2 = wb.create_sheet("IS")
        ws2["C3"] = "Hợp nhất"

    data = _xlsx_bytes(build)
    parsed = OpenpyxlExcelParser().parse_bytes(data)

    assert parsed.sheet_names == ("BS", "IS")
    assert parsed.parser_name == "openpyxl"
    assert len(parsed.source_sha256) == 64

    bs = parsed.sheets[0]
    by_addr = {cell.address: cell for cell in bs.cells}
    assert by_addr["A1"].value == "Doanh thu thuần"
    assert by_addr["A1"].row == 1
    assert by_addr["A1"].column == 1
    assert by_addr["A1"].sheet_name == "BS"
    assert by_addr["B2"].value == 1_234_567
    assert by_addr["B2"].value_text == "1234567"
    assert by_addr["A1"].value_kind == ExcelCellValueKind.SCALAR
    assert by_addr["B2"].value_kind == ExcelCellValueKind.SCALAR
    assert by_addr["B2"].formula_text is None
    assert by_addr["B2"].cached_value is None

    is_sheet = parsed.sheets[1]
    assert any(c.address == "C3" and c.value == "Hợp nhất" for c in is_sheet.cells)


def test_merged_cells_preserve_range() -> None:
    def build(wb: Workbook) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = "Merged"
        ws["A1"] = "Label"
        ws.merge_cells("A1:B2")

    parsed = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes(build))
    sheet = parsed.sheets[0]
    assert "A1:B2" in sheet.merged_ranges

    cells = {c.address: c for c in sheet.cells}
    assert cells["A1"].is_merged is True
    assert cells["A1"].merge_range == "A1:B2"
    assert cells["A1"].value == "Label"
    assert cells["A1"].value_kind == ExcelCellValueKind.SCALAR
    assert cells["B2"].is_merged is True
    assert cells["B2"].value is None
    assert cells["B2"].value_kind == ExcelCellValueKind.EMPTY
    assert cells["B2"].merge_range == "A1:B2"


def test_formulas_are_retained_not_evaluated() -> None:
    def build(wb: Workbook) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = "Calc"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"

    parsed = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes(build))
    cells = {c.address: c for c in parsed.sheets[0].cells}
    formula_cell = cells["A3"]
    assert formula_cell.value_kind == ExcelCellValueKind.FORMULA
    assert formula_cell.formula_text == "=SUM(A1:A2)"
    assert formula_cell.value == "=SUM(A1:A2)"
    assert formula_cell.value_text == "=SUM(A1:A2)"
    # Parser uses data_only=False — no calculated result is implied.
    assert formula_cell.cached_value is None
    assert formula_cell.value != 30


def test_hidden_and_empty_sheets_are_retained_with_visibility() -> None:
    def build(wb: Workbook) -> None:
        visible = wb.active
        assert visible is not None
        visible.title = "Visible"
        visible["A1"] = "shown"

        empty = wb.create_sheet("Empty")
        assert empty.max_row == 1

        hidden = wb.create_sheet("HiddenNotes")
        hidden["A1"] = "secret"
        hidden.sheet_state = "hidden"

        very_hidden = wb.create_sheet("VeryHidden")
        very_hidden["A1"] = "deeper"
        very_hidden.sheet_state = "veryHidden"

    parsed = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes(build))
    by_name = {sheet.name: sheet for sheet in parsed.sheets}
    assert set(by_name) == {"Visible", "Empty", "HiddenNotes", "VeryHidden"}

    assert by_name["Visible"].visibility == ExcelSheetVisibility.VISIBLE
    assert by_name["Empty"].visibility == ExcelSheetVisibility.VISIBLE
    assert by_name["Empty"].cells == ()
    assert by_name["HiddenNotes"].visibility == ExcelSheetVisibility.HIDDEN
    assert by_name["HiddenNotes"].cells[0].value == "secret"
    assert by_name["VeryHidden"].visibility == ExcelSheetVisibility.VERY_HIDDEN


def test_basic_number_formats_preserved() -> None:
    def build(wb: Workbook) -> None:
        ws = wb.active
        assert ws is not None
        ws["A1"] = 12.5
        ws["A1"].number_format = "0.00"
        ws["B1"] = date(2025, 12, 31)
        ws["B1"].number_format = "YYYY-MM-DD"

    parsed = OpenpyxlExcelParser().parse_bytes(_xlsx_bytes(build))
    cells = {c.address: c for c in parsed.sheets[0].cells}
    assert cells["A1"].number_format == "0.00"
    assert cells["A1"].value == pytest.approx(12.5)
    assert cells["B1"].number_format == "YYYY-MM-DD"
    assert cells["B1"].value == datetime(2025, 12, 31)
    assert cells["B1"].value_text == "2025-12-31T00:00:00"


def test_empty_bytes_raise() -> None:
    with pytest.raises(ExcelParseError, match="empty"):
        OpenpyxlExcelParser().parse_bytes(b"")


def test_corrupt_bytes_raise() -> None:
    with pytest.raises(ExcelParseError, match="Failed to open"):
        OpenpyxlExcelParser().parse_bytes(b"not-an-xlsx")


def test_missing_path_raises() -> None:
    with pytest.raises(FileNotFoundError):
        OpenpyxlExcelParser().parse_path(Path("does-not-exist.xlsx"))


@pytest.mark.asyncio
async def test_async_excel_parser_parse_bytes() -> None:
    def build(wb: Workbook) -> None:
        ws = wb.active
        assert ws is not None
        ws.title = "Async"
        ws["A1"] = "ok"

    parsed = await ExcelParser().parse_bytes(_xlsx_bytes(build))
    assert parsed.sheet_names == ("Async",)
    assert parsed.sheets[0].cells[0].address == "A1"
