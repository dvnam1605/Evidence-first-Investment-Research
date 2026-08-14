"""Unit tests for DOC-13 document processing pipeline."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import fitz
import pytest
from openpyxl import Workbook
from src.config.settings import ProcessingSettings
from src.domain.enums import BlockType, DetectedFileType, ProcessingStatus
from src.processing.pipeline import DocumentProcessor, DocumentProcessRequest
from src.processing.sections.models import StatementSection
from src.processing.tables import (
    ExtractedCell,
    ExtractedRow,
    ExtractedTable,
    PageTableExtractionIssue,
    TableExtractionContext,
    TableExtractionResult,
    TableExtractionStatus,
)
from src.processing.tables.raw_source import JsonRawTableSource
from src.services.process_factory import build_document_processor


def _pdf_bytes(text: str) -> bytes:
    doc = fitz.open()
    try:
        page = doc.new_page(width=595, height=842)
        page.insert_text((72, 72), text)
        return doc.tobytes()
    finally:
        doc.close()


def _xlsx_bytes() -> bytes:
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Bang can doi ke toan"
    ws["A1"] = "Tai san"
    ws["B1"] = 100
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_pipeline_processes_digital_pdf_without_ocr_engine() -> None:
    data = _pdf_bytes("CONSOLIDATED INCOME STATEMENT\nNet revenue 1,000")
    processor = DocumentProcessor(ocr_engine=None)
    result = await processor.process(
        DocumentProcessRequest(
            artifact_id=uuid4(),
            data=data,
            filename="fpt_income.pdf",
            title="Bao cao tai chinh hop nhat",
            document_type="financial_statement",
        )
    )
    assert result.file_type is DetectedFileType.PDF
    assert result.status in {ProcessingStatus.PROCESSED, ProcessingStatus.NEEDS_REVIEW}
    assert len(result.pages) == 1
    assert result.pages[0].text
    assert result.classification is not None
    assert result.error is None
    assert all(block.block_type in {BlockType.TEXT, BlockType.TABLE} for block in result.blocks)


async def test_pipeline_marks_unsupported_bytes_failed() -> None:
    processor = DocumentProcessor(ocr_engine=None)
    result = await processor.process(
        DocumentProcessRequest(
            artifact_id=uuid4(),
            data=b"not-a-real-document",
            filename="notes.txt",
        )
    )
    assert result.status is ProcessingStatus.FAILED
    assert result.file_type is DetectedFileType.UNKNOWN
    assert "unsupported_file_type" in result.warnings


async def test_pipeline_processes_xlsx() -> None:
    processor = DocumentProcessor(ocr_engine=None)
    result = await processor.process(
        DocumentProcessRequest(
            artifact_id=uuid4(),
            data=_xlsx_bytes(),
            filename="fs.xlsx",
            title="BCTC",
            document_type="financial_statement",
        )
    )
    assert result.file_type is DetectedFileType.XLSX
    assert result.status in {ProcessingStatus.PROCESSED, ProcessingStatus.NEEDS_REVIEW}
    assert len(result.pages) >= 1
    assert result.table_extraction is not None
    assert len(result.reconstructed_tables) >= 1
    assert any(block.block_type is BlockType.TABLE for block in result.blocks)


async def test_pipeline_flags_ocr_required_without_engine() -> None:
    # Below min_native_chars → detector requests OCR.
    data = _pdf_bytes("hi")
    result = await DocumentProcessor(ocr_engine=None).process(
        DocumentProcessRequest(artifact_id=uuid4(), data=data, filename="scan.pdf")
    )
    assert result.file_type is DetectedFileType.PDF
    assert result.ocr_decision is not None
    assert result.ocr_decision.needs_ocr is True
    assert result.status is ProcessingStatus.NEEDS_REVIEW
    assert "ocr_required_but_engine_not_configured" in result.warnings


async def test_pipeline_uses_sha_matched_vetted_raw_table_sidecar(tmp_path: Path) -> None:
    data = _pdf_bytes("raster placeholder")
    source_sha256 = hashlib.sha256(data).hexdigest()
    sidecar = tmp_path / "table.json"
    sidecar.write_text(
        json.dumps(
            {
                "source": {"sha256": source_sha256, "pdf": "statement.pdf"},
                "table": {
                    "page": 1,
                    "table_index": 0,
                    "table_id": "fixture:p1:t0",
                    "source_sha256": source_sha256,
                    "status": "OK",
                    "rows": [
                        {
                            "row_index": 0,
                            "cells": [
                                {
                                    "raw_text": "CONSOLIDATED INCOME STATEMENT",
                                    "row": 0,
                                    "column": 0,
                                    "page": 1,
                                    "table_id": "fixture:p1:t0",
                                }
                            ],
                        },
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "raw_text": "Revenue",
                                    "row": 1,
                                    "column": 0,
                                    "page": 1,
                                    "table_id": "fixture:p1:t0",
                                },
                                {
                                    "raw_text": "1000",
                                    "row": 1,
                                    "column": 1,
                                    "page": 1,
                                    "table_id": "fixture:p1:t0",
                                },
                            ],
                        },
                    ],
                },
            }
        ),
        encoding="utf-8",
    )
    artifact_id = uuid4()
    processor = build_document_processor(
        SimpleNamespace(
            processing=ProcessingSettings(raw_table_dir=tmp_path, _env_file=None)
        ),
        use_optional_ocr=False,
    )
    result = await processor.process(
        DocumentProcessRequest(
            artifact_id=artifact_id,
            data=data,
            filename="statement.pdf",
        )
    )

    assert result.table_extraction is not None
    assert len(result.reconstructed_tables) == 1
    raw = result.reconstructed_tables[0].raw
    assert raw.document_id == artifact_id
    assert raw.artifact_id == artifact_id
    assert raw.table_id == f"{artifact_id}:p1:raw:t0"
    assert raw.extractor_name == "json_raw_table"
    assert raw.rows[1].cells[1].raw_text == "1000"
    assert result.sections is not None
    assert any(hit.section is StatementSection.INCOME_STATEMENT for hit in result.sections.hits)
    table_block = next(block for block in result.blocks if block.block_type is BlockType.TABLE)
    payload = table_block.content["intermediate"]
    assert isinstance(payload, dict)
    raw_payload = payload["raw"]
    assert isinstance(raw_payload, dict)
    assert raw_payload["document_id"] == str(artifact_id)
    assert raw_payload["source_sha256"] == source_sha256
    cells = raw_payload["cells"]
    assert isinstance(cells, list)
    assert cells[0]["page"] == 1
    assert cells[0]["table_id"] == f"{artifact_id}:p1:raw:t0"

    mismatched = await processor.process(
        DocumentProcessRequest(
            artifact_id=uuid4(),
            data=_pdf_bytes("different source document"),
            filename="different.pdf",
        )
    )
    assert len(mismatched.reconstructed_tables) == 0


async def test_pipeline_processes_real_fpt_raw_statement_sidecars() -> None:
    root = Path(__file__).resolve().parents[2] / "fixtures" / "processing" / "tables"
    pdf_path = root / "fpt_real" / (
        "20260727_FPT_Consolidated_Financial_Statements_for_Q22026_"
        "055fe1ecd3_024b14f8ceff.pdf"
    )
    if not pdf_path.is_file():
        pytest.skip("local real FPT PDF fixture is unavailable")
    raw_dir = root / "fpt_raw"
    source = JsonRawTableSource(
        tuple(sorted(raw_dir.glob("q2_2026_consol_*.json")))
    )
    artifact_id = uuid4()
    result = await DocumentProcessor(
        ocr_engine=None,
        raw_table_source=source,
    ).process(
        DocumentProcessRequest(
            artifact_id=artifact_id,
            data=pdf_path.read_bytes(),
            filename=pdf_path.name,
        )
    )

    assert len(result.reconstructed_tables) == 3
    assert result.sections is not None
    assert {hit.section for hit in result.sections.hits} >= {
        StatementSection.BALANCE_SHEET,
        StatementSection.INCOME_STATEMENT,
        StatementSection.CASH_FLOW_STATEMENT,
    }
    assert [table.raw.page for table in result.reconstructed_tables] == [3, 7, 10]
    assert all(table.raw.document_id == artifact_id for table in result.reconstructed_tables)


async def test_pipeline_produces_reviewable_tables_for_five_real_fpt_reports() -> None:
    root = Path(__file__).resolve().parents[2] / "fixtures" / "processing" / "tables"
    real_dir = root / "fpt_real"
    documents = {
        "20260319_fpt_bctc_hop_nhat_nam_2025_da_kiem_toan_0159f59257_630f61f6ef9f.pdf": (
            8,
            "TÀI SẢN NGẮN HẠN",
            "58.137.438.254.908",
            "45.535.942.846.453",
        ),
        "20260319_fpt_bctc_rieng_nam_2025_da_kiem_toan_f67c6c805b_4fce6b21101d.pdf": (
            8,
            "TÀI SẢN NGẮN HẠN",
            "13.998.282.834.060",
            "12.764.466.007.896",
        ),
        "20260424_fpt_bctc_hop_nhat_quy_1_nam_2026_d03fbbebf9_5adaee07d5aa.pdf": (
            3,
            "TÀI SẢN NGẮN HẠN",
            "41.527.873.060.120",
            "58.137.438.254.908",
        ),
        "20260424_fpt_bctc_rieng_quy_1_nam_2026_4bcdf6ea7d_11c9709914cd.pdf": (
            4,
            "TÀI SẢN NGẮN HẠN",
            "10.626.141.721.085",
            "13.998.282.834.060",
        ),
        "20260727_FPT_Separarate_Financial_Statements_for_Q22026_494daa3546_6034d952ce36.pdf": (
            4,
            "CURRENT ASSETS",
            "10.509.464.631.417",
            "13.998.282.834.060",
        ),
    }
    paths = {filename: real_dir / filename for filename in documents}
    if not all(path.is_file() for path in paths.values()):
        pytest.skip("local real FPT PDF fixtures are unavailable")
    processor = build_document_processor(use_optional_ocr=False)

    for filename, (page_number, label, current_value, comparative_value) in documents.items():
        artifact_id = uuid4()
        result = await processor.process(
            DocumentProcessRequest(
                artifact_id=artifact_id,
                data=paths[filename].read_bytes(),
                filename=filename,
            )
        )

        assert result.status is ProcessingStatus.NEEDS_REVIEW
        assert result.sections is not None
        assert any(
            hit.section is StatementSection.BALANCE_SHEET
            and hit.page_number == page_number
            for hit in result.sections.hits
        )
        table = next(
            table
            for table in result.reconstructed_tables
            if table.raw.page == page_number
        )
        assert table.raw.document_id == artifact_id
        assert table.raw.source_sha256
        row_values = [cell.raw_text for cell in table.raw.rows[2].cells]
        assert label in row_values
        assert row_values[-2:] == [current_value, comparative_value]
        assert table.raw.rows[2].cells[-1].bbox is not None
        assert table.raw.rows[2].cells[-1].bbox_estimated is True
        assert table.raw.status.value == "NEEDS_REVIEW"
        assert table.table_type.value == "balance_sheet"


async def test_pipeline_surfaces_malformed_raw_table_sidecar(tmp_path: Path) -> None:
    (tmp_path / "broken.json").write_text("{not valid json", encoding="utf-8")
    processor = build_document_processor(
        SimpleNamespace(
            processing=ProcessingSettings(raw_table_dir=tmp_path, _env_file=None)
        ),
        use_optional_ocr=False,
    )

    result = await processor.process(
        DocumentProcessRequest(
            artifact_id=uuid4(),
            data=_pdf_bytes("native text"),
            filename="statement.pdf",
        )
    )

    assert result.status is ProcessingStatus.NEEDS_REVIEW
    assert any(
        warning.startswith("raw_table_sidecar_failed:") for warning in result.warnings
    )


async def test_matching_sidecar_survives_unrelated_malformed_file(tmp_path: Path) -> None:
    data = _pdf_bytes("statement")
    source_sha256 = hashlib.sha256(data).hexdigest()
    (tmp_path / "broken.json").write_text("{broken", encoding="utf-8")
    (tmp_path / "valid.json").write_text(
        json.dumps(
            {
                "source": {"sha256": source_sha256},
                "table": {
                    "page": 1,
                    "table_id": "fixture:p1:t0",
                    "matrix": [["INCOME STATEMENT"], ["Revenue", "1000"]],
                },
            }
        ),
        encoding="utf-8",
    )
    processor = build_document_processor(
        SimpleNamespace(
            processing=ProcessingSettings(raw_table_dir=tmp_path, _env_file=None)
        ),
        use_optional_ocr=False,
    )

    result = await processor.process(
        DocumentProcessRequest(artifact_id=uuid4(), data=data, filename="statement.pdf")
    )

    assert len(result.reconstructed_tables) == 1
    assert result.reconstructed_tables[0].raw.rows[1].cells[1].raw_text == "1000"
    assert any(
        warning.startswith("raw_table_sidecar_failed:broken.json")
        for warning in result.warnings
    )


async def test_partial_sidecar_preserves_native_table_and_page_issue() -> None:
    data = _pdf_bytes("statement")
    artifact_id = uuid4()

    class NativeTables:
        def extract_pdf(
            self,
            _: bytes,
            *,
            context: TableExtractionContext,
        ) -> TableExtractionResult:
            table_id = f"{artifact_id}:p1:native:t0"
            return TableExtractionResult(
                tables=(
                    ExtractedTable(
                        page=1,
                        table_id=table_id,
                        rows=(
                            ExtractedRow(
                                row_index=0,
                                cells=(
                                    ExtractedCell(
                                        raw_text="Native table",
                                        row=0,
                                        column=0,
                                        page=1,
                                        table_id=table_id,
                                    ),
                                ),
                            ),
                        ),
                        document_id=artifact_id,
                        status=TableExtractionStatus.OK,
                    ),
                ),
                page_issues=(
                    PageTableExtractionIssue(
                        page=1,
                        status=TableExtractionStatus.NEEDS_REVIEW,
                        reason="native_issue",
                    ),
                ),
                context=context,
            )

    class PartialRawSource:
        @property
        def name(self) -> str:
            return "partial"

        def load_for(
            self,
            *,
            source_sha256: str,
            context: TableExtractionContext,
        ) -> TableExtractionResult:
            table_id = f"{artifact_id}:p1:raw:t0"
            return TableExtractionResult(
                tables=(
                    ExtractedTable(
                        page=1,
                        table_id=table_id,
                        table_index=0,
                        rows=(
                            ExtractedRow(
                                row_index=0,
                                cells=(
                                    ExtractedCell(
                                        raw_text="Partial raw table",
                                        row=0,
                                        column=0,
                                        page=1,
                                        table_id=table_id,
                                    ),
                                ),
                            ),
                        ),
                        document_id=artifact_id,
                        source_sha256=source_sha256,
                        warnings=("partial_statement_region",),
                    ),
                ),
                context=context,
            )

    result = await DocumentProcessor(
        tables=NativeTables(),  # type: ignore[arg-type]
        raw_table_source=PartialRawSource(),
        ocr_engine=None,
    ).process(
        DocumentProcessRequest(
            artifact_id=artifact_id,
            data=data,
            filename="statement.pdf",
        )
    )

    assert {table.raw.table_id for table in result.reconstructed_tables} == {
        f"{artifact_id}:p1:native:t0",
        f"{artifact_id}:p1:raw:t0",
    }
    assert result.table_extraction is not None
    assert [issue.reason for issue in result.table_extraction.page_issues] == [
        "native_issue"
    ]
